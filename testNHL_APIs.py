import requests
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime, timedelta, date
from PIL import Image

# from constants import proTeamTuple

#A class to find a team by the abbrev or full name and return the tuple of info for use later
class ProTeams:
    def __init__(self):
        # self.proTeamInfo = {}
        self.proTeamTuple = [("BOS","Boston Bruins","BOS.png"),
            ("BUF","Buffalo Sabres","BUF.png"),
            ("CGY","Calgary Flames","CGY.png"),
            ("CHI","Chicago Blackhawks","CHI.png"),
            ("DET","Detroit Red Wings","DET.png"),
            ("EDM","Edmonton Oilers","EDM.png"),
            ("CAR","Carolina Hurricanes","CAR.png"),
            ("LAK","Los Angeles Kings","LA.png"),
            ("DAL","Dallas Stars","DAL.png"),
            ("MTL","Montréal Canadiens","MTL.png"),
            ("NJD","New Jersey Devils","NJ.png"),
            ("NYI","New York Islanders","NYI.png"),
            ("NYR","New York Rangers","NYR.png"),
            ("OTT","Ottawa Senators","OTT.png"),
            ("PHI","Philadelphia Flyers","PHI.png"),
            ("PIT","Pittsburgh Penguins","PIT.png"),
            ("COL","Colorado Avalanche","COL.png"),
            ("SJS","San Jose Sharks","SJ.png"),
            ("STL","St. Louis Blues","STL.png"),
            ("TBL","Tampa Bay Lightning","TB.png"),
            ("TOR","Toronto Maple Leafs","TOR.png"),
            ("VAN","Vancouver Canucks","VAN.png"),
            ("WSH","Washington Capitals","WSH.png"),
            ("ARI","Arizona Coyotes","ARI.png"),
            ("ANA","Anaheim Ducks","ANA.png"),
            ("FLA","Florida Panthers","FLA.png"),
            ("NSH","Nashville Predators","NSH.png"),
            ("WPG","Winnipeg Jets","WPG.png"),
            ("CBJ","Columbus Blue Jackets","CBJ.png"),
            ("MIN","Minnesota Wild","MIN.png"),
            ("VGK","Vegas Golden Knights","VGK.png"),
            ("SEA","Seattle Kraken","SEA.png")
        ]

    def findTeamRowInTuple (self, teamName):
        #If abbrev is mixed case, check for a length of 3 (3=abbrev), then make value all uppercase
        if len(teamName) == 3:
            teamName=teamName.upper()
        for sublist in self.proTeamTuple:
            for element in sublist:
                if element == teamName:
                    return sublist
        return None #if value not found
    
    def findIndexOfTeamInTuple (self, teamName):
        if len(teamName) == 3:
            teamName=teamName.upper()
        for sublist in self.proTeamTuple:
            for element in sublist:
                if element == teamName:
                    teamIndex=self.proTeamTuple.index(sublist)
                    return teamIndex
        return None #if value not found

#This class parses thru the API json and creates a text file of the info for the schedule
class GetNHLSchedule:
    def __init__(self,weekDate):
        weekDateCall='https://api-web.nhle.com/v1/schedule/'+weekDate #Looks like: 'https://api-web.nhle.com/v1/schedule/2024-03-25'
        self.nhlApi=requests.get(weekDateCall)

    def delete_file_if_exists(self, fname): #check for file, delete if exists to avoid duplicate schedules
        if os.path.exists(fname): # Check if the file exists
            os.remove(fname) # If it exists, delete the file
            print(f"File '{fname}' deleted.")
        else:
            print(f"File '{fname}' does not exist.")

    def getNhlGameInfo(self, fname):
        r=self.nhlApi
        if r.status_code != 200:
            print ("Problem connecting with NHL API")
            exit
        self.delete_file_if_exists(fname) #delete the file if it already exists so we don't append to it
        f = open(fname, "a")
        theJSON = json.loads(r.content)
        for i in theJSON["gameWeek"]:
            dateGameOfWeek = i["date"]
            dayAbbrev = i["dayAbbrev"]
            dayNumberOfGames = i["numberOfGames"]
            for aRow in i["games"]:
                awayTeamName = aRow["awayTeam"]["placeName"]["default"]
                awayTeamAbbrev = aRow["awayTeam"]["abbrev"]
                homeTeamName = aRow["homeTeam"]["placeName"]["default"]
                homeTeamAbbrev = aRow["homeTeam"]["abbrev"]
                gameInfo=dateGameOfWeek+","+dayAbbrev+","+str(dayNumberOfGames)+","+awayTeamAbbrev+","+homeTeamAbbrev+'\n'
                f.write(gameInfo)
        f.close()

#A class to create the excel file with the scheduled data
class WriteNHLSchedule:
    def __init__(self, xName):
        self.filename = xName
        self.workbook = Workbook()
        self.ws = self.workbook.active

    def set_row_height(self, row, height):
        self.ws.row_dimensions[row].height = height

    def set_column_width(self, column, width):
        self.ws.column_dimensions[column].width = width

    def set_cell_font(self, row, column, font_name='Arial', font_size=11, bold=False, color=None):
        cell = self.ws.cell(row=row, column=column)
        cell.font = Font(name=font_name, size=font_size, bold=bold, color=color)

    def set_cell_alignment(self, row, column, horizontal='center', vertical='center'):
        cell = self.ws.cell(row=row, column=column)
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    def set_cell_border(self, row, column, border_style='thin'):
        cell = self.ws.cell(row=row, column=column)
        border = Border(left=Side(style=border_style), right=Side(style=border_style), 
                        top=Side(style=border_style), bottom=Side(style=border_style))
        cell.border = border

    def set_cell_fill_color(self, row, column, color='FFFFFF'):
        cell = self.ws.cell(row=row, column=column)
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.fill = fill

    def set_column_width(self, start_column, end_column, width):
        for col in range(start_column, end_column):
            self.ws.column_dimensions[self.ws.cell(row=1, column=col).column_letter].width = width

    def write_row_data(self, row, data):
        for i, value in enumerate(data, start=1):
            self.ws.cell(row=row, column=i, value=value)

    def write_column_data(self, row, column, value):
        self.ws.cell(row=row, column=column, value=value)

    def insert_team_logo(self, row, column, imageName):
        row=str(row)
        cellName=column+row
        from PIL import Image as PILImage #Pillow (PIL) was needed for dealing with images
        pil_image = PILImage.open(imageName)
        excel_image = ExcelImage(pil_image)
        self.ws.add_image(excel_image, cellName)

    def save_excel(self):
        self.workbook.save(self.filename)

    def close_excel(self):
        self.workbook.close(self.filename)

#Create a dictionary that will keep track of a team's game count for the week
class TeamGameCount:
    def __init__(self):
        self.teamListCount = {
            "BOS": 0,
            "BUF": 0,
            "CGY": 0,
            "CHI": 0,
            "DET": 0,
            "EDM": 0,
            "CAR": 0,
            "LAK": 0,
            "DAL": 0,
            "MTL": 0,
            "NJD": 0,
            "NYI": 0,
            "NYR": 0,
            "OTT": 0,
            "PHI": 0,
            "PIT": 0,
            "COL": 0,
            "SJS": 0,
            "STL": 0,
            "TBL": 0,
            "TOR": 0,
            "VAN": 0,
            "WSH": 0,
            "ARI": 0,
            "ANA": 0,
            "FLA": 0,
            "NSH": 0,
            "WPG": 0,
            "CBJ": 0,
            "MIN": 0,
            "VGK": 0,
            "SEA": 0
        }

    def teamGameCountIncrement(self, key):
        if key in self.teamListCount:
            self.teamListCount[key] += 1
        else:
            print(f"Key '{key}' not found.")

    def teamGameCountRetrieval(self, key):
        return self.teamListCount.get(key)

#######
#Beginning of main code
team=ProTeams() #Class to find a team by the abbrev or full name and return the tuple of info

#Calling a class to parse thru the API json and creates a text file of the info for the schedule
dateForTheWeek='2024-04-08'
a=GetNHLSchedule(dateForTheWeek)
filename = "c:\\Temp\\demoHockeySchedule.txt"
a.getNhlGameInfo(filename)

#Calling a class to create the excel file with the scheduled data
xName="c:\\Temp\\hockeydemo.xlsx"
daysOfWeek = ["Abbrev","Team Logo","MON","TUE","WED","THU","FRI","SAT","SUN","Game Count"]
daysOfWeekCount=len(daysOfWeek)
loopPlusThree=daysOfWeekCount+3
excelNhlSchedule=WriteNHLSchedule(xName)
excelNhlSchedule.set_row_height(1, 15)
excelNhlSchedule.set_column_width(1, 10, 12)  # Set columns A to H to width 10
excelNhlSchedule.set_column_width(10, 11, 15)
loopRange=loopPlusThree+1
for i in range(loopRange):
    i+=1
    excelNhlSchedule.set_cell_font(1, i, bold=True, color='090DF8')  # Set font bold and blue color for cell A1
    excelNhlSchedule.set_cell_alignment(1, i, horizontal='center', vertical='center')  # Center align cell A1
    excelNhlSchedule.set_cell_border(1, i)  # Add thin border to cell A1
    excelNhlSchedule.set_cell_fill_color(1, i, color='EEF8A6')  # Set light yellow fill color for cell A1

dayOfWeekIndex=0
excelNhlSchedule.write_row_data(1, daysOfWeek)  # Write data to row 1
 
#get each day (date) for the week, for the column header
daysForTheWeek=[]
for i in range(8):
    if i==0:
        daysForTheWeek=[dateForTheWeek]
    else:
        Begindate = date.fromisoformat(dateForTheWeek)
        incremented_date = Begindate + timedelta(days=i)
        daysForTheWeek.append(incremented_date)
for j in range(11):
    colNbr=j+1
    excelNhlSchedule.set_cell_alignment(2, colNbr, horizontal='center', vertical='center')
    excelNhlSchedule.set_cell_font(2, colNbr, bold=True, color='1fa180')

dateHeader=[" "," ",daysForTheWeek[0],daysForTheWeek[1],daysForTheWeek[2],daysForTheWeek[3],daysForTheWeek[4],daysForTheWeek[5],daysForTheWeek[6],""]
excelNhlSchedule.write_row_data(2, dateHeader)

#Setup left columns for NHL teams
imagePath="C:\\Temp\\HockeyTeamLogos\\"
row=4
for i, value in enumerate(team.proTeamTuple, start=00):
    excelNhlSchedule.set_row_height(row, 35)
    excelNhlSchedule.set_cell_alignment(row, 1, horizontal='center', vertical='center')
    excelNhlSchedule.set_cell_font(row, 1, font_name="Georgia", bold=True, color='367ee7')
    excelNhlSchedule.write_column_data(row, 1, value[0]) #Can use numbers too for column name
    imageName=imagePath+value[2]
    excelNhlSchedule.insert_team_logo(row, "B", imageName)
    row+=1

#Cell color fill for the empty cells not used
cellFill=[1,2,10]
for x in range(2, 4):
    for i, value in enumerate(cellFill, start=0):
        excelNhlSchedule.set_cell_fill_color(x, cellFill[i], color='c0c0c0')

#Write the schedule now. Find the matching team and column to write to
#Look up the teams and fill it in on the spreadsheet - basically twice for each team
teamGameCnt = TeamGameCount() #used to keep track of total games for the week by team
rowOffset=4 #to get positioned, first team in spreadsheet starts at row 4
col=2
colLetter='B'
firstRetrieval="True"
f = open(filename,'r')
i=0
while True:
        x=f.readline()
        if not x:
            print("EOF reached")
            f.close()
            break
        #Parse line: date, day, nbr of games, away, home
        gameInfo= x.split(",") #2024-03-25,MON,2,VGK,STL
        #These if statements are to control the logic of what column to write to
        if firstRetrieval=="True":
            firstRetrieval="False"
            currentGameDate=gameInfo[0]
            nextGameDate=currentGameDate
            i+=1
            col+=1
            colLetter=chr (ord (colLetter) + 1) #columns are letters, increment the column to write to the correct one starting with C
            excelNhlSchedule.set_cell_font(3, col, bold=True, color='17020e')
            excelNhlSchedule.set_cell_alignment(3, col, horizontal='center', vertical='center')
            excelNhlSchedule.set_cell_fill_color(3, col, color='0fcbfd') 
            excelNhlSchedule.write_column_data(3, col, gameInfo[2])
        else:
            if i<=int(gameInfo[2]):
                i+=1
                nextGameDate=gameInfo[0]
        
        if currentGameDate!=nextGameDate:
            col+=1
            colLetter=chr (ord (colLetter) + 1) #columns are letters, increment the column to write to the correct one starting with C

        #Get info setup 
        awayTeam=gameInfo[3]
        homeTeam=gameInfo[4]
        homeTeam=homeTeam[0:3] #returning the first 3 characters to avoid the \n
        awayTeamRowInfo=team.findTeamRowInTuple(awayTeam) #team name abbrev in, returns "BOS","Boston Bruins","BOS.png"
        if awayTeamRowInfo == None: #exit when no team was found; probably a typo
            print("No Away team was found, exiting program as there is a problem; possibly a typo on team name.")
            sys.exit()
        homeTeamRowInfo=team.findTeamRowInTuple(homeTeam) #team name abbrev in, returns "BOS","Boston Bruins","BOS.png"
        if homeTeamRowInfo == None: #exit when no team was found; probably a typo
            print("No Home team was found, exiting program as there is a problem; possibly a typo on team name.")
            sys.exit()
        #Process the matchups - first the Away Team appears in the list
        #Find the Away Team POSITION in the spreadsheet, then write/insert the HOME Team logo
        #first team is always Away Team gameInfo[3]
        indexOfTeam=team.findIndexOfTeamInTuple(awayTeam) #using this to figure out where in the spreadsheet the teams are located
        rowPosition=indexOfTeam+rowOffset
        excelNhlSchedule.set_cell_alignment(rowPosition, col, horizontal='center', vertical='center')
        excelNhlSchedule.set_cell_fill_color(rowPosition, col, color='4edc9c')
        #Now insert the Home Team image
        imageName=imagePath+homeTeamRowInfo[2] #name of team image
        excelNhlSchedule.insert_team_logo(rowPosition, colLetter, imageName)
        teamGameCnt.teamGameCountIncrement(awayTeam) 
        #Now doing the other combo of the matchup
        indexOfTeam=team.findIndexOfTeamInTuple(homeTeam) #using this to figure out where in the spreadsheet the teams are located
        rowPosition=indexOfTeam+rowOffset
        excelNhlSchedule.set_cell_alignment(rowPosition, col, horizontal='center', vertical='center')
        excelNhlSchedule.set_cell_fill_color(rowPosition, col, color='ff1919')
        #Now insert the Away Team image
        imageName=imagePath+awayTeamRowInfo[2] #name of team image
        excelNhlSchedule.insert_team_logo(rowPosition, colLetter, imageName)
        teamGameCnt.teamGameCountIncrement(homeTeam)
        if i==int(gameInfo[2]):
            i=0
            firstRetrieval="True"
        elif i>int(gameInfo[2]):
            sys.exit("Problem with date count/comparison, got larger, exiting out.")

for i, value in enumerate(team.proTeamTuple, start=00):
    indexOfTeam=team.findIndexOfTeamInTuple(value[0]) #using this to figure out where in the spreadsheet the teams are located
    rowPosition=indexOfTeam+rowOffset
    gameCount=teamGameCnt.teamGameCountRetrieval(value[0])
    excelNhlSchedule.set_cell_font(rowPosition, 10, bold=True, color='17020e')
    excelNhlSchedule.set_cell_alignment(rowPosition, 10, horizontal='center', vertical='center')
    excelNhlSchedule.set_cell_fill_color(rowPosition, 10, color='0fcbfd')
    excelNhlSchedule.write_column_data(rowPosition, 10, gameCount)

excelNhlSchedule.save_excel()

# if __name__ == "__main__":
#     main()