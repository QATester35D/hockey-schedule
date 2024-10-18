from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage

#A class to create the excel file with the scheduled data
class WriteNHLSchedule:
    def __init__(self, xName):
        self.filename = xName
        self.workbook = Workbook()
        self.ws = self.workbook.active

    # def create_worksheet(self, title):
    #     sheet = self.workbook
    #     sheet2 = sheet.create_sheet(title)

    def set_row_height(self, row, height):
        self.ws.row_dimensions[row].height = height

    def set_column_width(self, column, width):
        self.ws.column_dimensions[column].width = width

    def set_cell_font(self, row, column, font_name='Arial', font_size=11, bold=False, color=None):
        cell = self.ws.cell(row=row, column=column)
        cell.font = Font(name=font_name, size=font_size, bold=bold, color=color)

    def set_cell_alignment(self, row, column, horizontal='center', vertical='center'):
        cell = self.ws.cell(row=row, column=column)
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

    def set_cell_border(self, row, column, border_style='thin'):
        cell = self.ws.cell(row=row, column=column)
        border = Border(left=Side(style=border_style), right=Side(style=border_style), 
                        top=Side(style=border_style), bottom=Side(style=border_style))
        cell.border = border

    def set_cell_fill_color(self, row, column, color='FFFFFF'):
        cell = self.ws.cell(row=row, column=column)
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.fill = fill

    def set_column_width(self, start_column, end_column, width):
        for col in range(start_column, end_column):
            self.ws.column_dimensions[self.ws.cell(row=1, column=col).column_letter].width = width

    def write_row_data(self, row, data):
        for i, value in enumerate(data, start=1):
            self.ws.cell(row=row, column=i, value=value)

    def write_column_data(self, row, column, value):
        self.ws.cell(row=row, column=column, value=value)

    def insert_team_logo(self, row, column, imageName):
        row=str(row)
        cellName=column+row
        from PIL import Image as PILImage #Pillow (PIL) was needed for dealing with images
        pil_image = PILImage.open(imageName)
        excel_image = ExcelImage(pil_image)
        self.ws.add_image(excel_image, cellName)

    def save_excel(self):
        self.workbook.save(self.filename)

    def close_excel(self):
        self.workbook.close(self.filename)